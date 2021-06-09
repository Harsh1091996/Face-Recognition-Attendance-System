from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell import Cell
import time
import os
import sqlite3
import face_recognition
from PIL import Image,ImageDraw 
import numpy as np
import picamera

def getDateColumn():
    currentDate = time.strftime("%d_%m_%y")
    wb = load_workbook(filename = "reports.xlsx")
    sheet = wb['Cse52']
    for i in range(1, sheet.max_row+1):
        col = get_column_letter(i)
        if sheet[str(col)+'1'].value == currentDate:
            return col
#database connection
def takeattendance():
 camera=picamera.PiCamera()
 camera.capture('/home/pi/webapp/pics/all.jpg')
 conn = sqlite3.connect('Face-DataBase')
 c = conn.cursor()

#get current date
 currentDate = time.strftime("%d_%m_%y")

#create a workbook and add a worksheet
 if(os.path.exists('./reports.xlsx')):
    wb = load_workbook(filename = "reports.xlsx")
    sheet = wb['Cse52']
    # sheet[ord() + '1']
    for col_index in range(1, 100):
        col = get_column_letter(col_index)
        if sheet['%s%s' % (col,1)].value is None:
            col2 = get_column_letter(col_index - 1)
            # print sheet.cell('%s%s'% (col2, 1)).value
            if sheet['%s%s' % (col2,1)] != currentDate:
                sheet['%s%s' % (col,1)] = currentDate
            break

    #saving the file
    wb.save(filename = "reports.xlsx")
        
 else:
    wb = Workbook()
    dest_filename = 'reports.xlsx'
    c.execute("SELECT * FROM CSE52 ORDER BY Roll ASC")
    
    #creating worksheet and giving names to column
    ws1 = wb.active
    ws1.title = "Cse52"
    ws1.append(('Roll Number', 'Name', currentDate))
    ws1.append(('', '', ''))

    #entering students information from database
    while True:
        a = c.fetchone()
        if a == None:
            break
        else:
            ws1.append((a[2], a[1]))

    #saving the file
    wb.save(filename = dest_filename)




#get current date
 currentDate = time.strftime("%d_%m_%y")
 wb = load_workbook(filename = "reports.xlsx")
 sheet = wb['Cse52']


#face_recognition


 attend = [0 for i in range(100)]

 c.execute("SELECT * FROM CSE52 ORDER BY Roll ASC")
 kfe=[]
 kfname=[]
 while True:
        a = c.fetchone()
        if a == None:
            break
        else:
            kfname.append(a[2])
            k=np.fromstring(a[3], dtype=float,sep=' ')
            kfe.append(k)
            
        
 unknownimg=face_recognition.load_image_file("/home/pi/webapp/pics/all.jpg")
 fls=face_recognition.face_locations(unknownimg)
 face_encodings=face_recognition.face_encodings(unknownimg,fls)
 pil_img=Image.fromarray(unknownimg)
 draw=ImageDraw.Draw(pil_img)
 for (t,r,b,l),face_encoding in zip(fls,face_encodings):
    match=face_recognition.compare_faces(kfe,face_encoding)
    name="unknown"
    if True in match:
                first_match_index=match.index(True)
                name=kfname[first_match_index]
                c.execute("SELECT * FROM CSE52 WHERE Roll= ? ", (name,))
                row = c.fetchone()
                attend[int(row[0])] += 1

 for row in range(2, sheet.max_row+1):
    rn = sheet['A'+str(row)].value
    if rn is not None:
        rn = rn[-2:]
        if attend[int(rn)] != 0:
            col = getDateColumn()
            sheet['%s%s' % (col, str(row))].value = 1

 wb.save(filename = "reports.xlsx")
 camera.close()        
